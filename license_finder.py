import requests
from openpyxl import load_workbook
from openpyxl import Workbook
wb = load_workbook('TPS_Declaration_Sheet_SandAccumulation.xlsx')
ws = wb["Third party Software List"]
new_workbook = Workbook()
sheet = new_workbook.active

i=1          
for row in ws.iter_rows():
    i=i+1
    name=str(row[1].value)
    ver = str(row[2].value)
    if row[1].value !=None and row[2].value != None:
        url = "https://api.clearlydefined.io/definitions/pypi/pypi/-/"+name+"/"+ver
        print(url)
        payload={}
        headers = {
          'accept': 'application/json'
        }

        response = requests.request("GET", url, headers=headers, data=payload)
        if response.ok:
            res = response.json()
            lic=res.get("licensed").get("declared")
            print(lic)
            sheet.cell(row=i, column=1).value = name 
            sheet.cell(row=i, column=2).value = ver
            sheet.cell(row=i, column=3).value = lic

    new_workbook.save(filename="licenses.xlsx")
              
                
